import os
import tempfile
import logging
import pandas as pd
from io import BytesIO
from fpdf import FPDF
from app.db.supabase_client import get_supabase
from google import genai

logger = logging.getLogger(__name__)

def _get_ai_summary(items: list, stock_value: float, total_revenue: float, total_profit: float, top_item: str) -> str:
    """Shared helper: generate Gemini executive summary focusing on P&L."""
    if not items:
        return "Склад пуст. Нет данных для анализа."
    prompt = (
        f"Контекст: Складской остаток (P&L отчет). "
        f"Текущая стоимость остатков: {stock_value} сом. "
        f"Общая выручка: {total_revenue} сом. "
        f"Чистая прибыль: {total_profit} сом. "
        f"Самый продаваемый товар: {top_item}. "
        "Составь профессиональное финансовое резюме (Executive Summary) для CFO. "
        "ОБЯЗАТЕЛЬНЫЕ ПРАВИЛА (СТРОГО): "
        "1) ЗАПРЕЩЕНО использовать любые Markdown-символы (никаких звездочек **, *, #, жирного текста или курсива). Пиши просто обычный текст. "
        "2) Валюта ТОЛЬКО 'сом'. Запрещены упоминания рублей, долларов или других валют. "
        "3) Включи в текст фразу: 'Общая выручка составила X сом, чистая прибыль Y сом. Самый продаваемый товар — Z', подставив цифры. "
        "Сделай текст на 3-4 предложения, четким, сухим финансовым языком."
    )
    try:
        client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
        response = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
        text = response.text.replace("\n", " ").strip()
        # Clean up any potential markdown traces the LLM might have hallucinated
        return text.replace("*", "").replace("#", "")
    except Exception as e:
        logger.error(f"AI summary error: {e}")
        return "Не удалось сгенерировать Финансовый AI-отчет. Ошибка подключения к Gemini."


def generate_inventory_pdf() -> str:
    db = get_supabase()
    res = db.table("inventory").select("*").execute()
    items = res.data or []
    stock_value = sum((i.get("quantity", 0) or 0) * (i.get("price", 0.0) or 0.0) for i in items)
    
    total_revenue = sum((i.get("sold_quantity", 0) or 0) * (i.get("price", 0.0) or 0.0) for i in items)
    total_profit = sum(((i.get("price", 0.0) or 0.0) - (i.get("cost_price", 0.0) or 0.0)) * (i.get("sold_quantity", 0) or 0) for i in items)
    
    sorted_by_sales = sorted(items, key=lambda x: x.get("sold_quantity", 0) or 0, reverse=True)
    top_item = sorted_by_sales[0].get("name", "Нет") if sorted_by_sales else "Нет"

    summary = _get_ai_summary(items, stock_value, total_revenue, total_profit, top_item)

    pdf = FPDF()
    pdf.add_page()

    font_path = os.path.join(os.getcwd(), "backend", "Roboto-Regular.ttf")
    if not os.path.exists(font_path) and os.path.exists("Roboto-Regular.ttf"):
        font_path = "Roboto-Regular.ttf"

    font_loaded = False
    try:
        if os.path.exists(font_path):
            pdf.add_font("Roboto", "", font_path, uni=True)
            pdf.set_font("Roboto", "", 14)
            font_loaded = True
        else:
            pdf.set_font("Helvetica", "", 14)
    except Exception:
        pdf.set_font("Helvetica", "", 14)

    pdf.cell(200, 10, txt="P&L Финансовый Отчет (Inventory Report)", ln=True, align="C")
    pdf.ln(10)
    if font_loaded:
        pdf.set_font("Roboto", "", 11)
    pdf.multi_cell(0, 10, txt=f"AI Финансовая Сводка: {summary}")
    pdf.ln(10)
    pdf.cell(0, 10, txt=f"Общая выручка: {total_revenue} сом | Прибыль: {total_profit} сом", ln=True)
    pdf.ln(5)
    for idx, item in enumerate(items, 1):
        name = item.get("name", "Unknown")
        qty = item.get("quantity", 0)
        price = item.get("price", 0.0)
        pdf.cell(0, 8, txt=f"{idx}. {name} | Остаток: {qty} | Цена продажи: {price}", ln=True)

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf.output(temp_file.name)
    return temp_file.name


def generate_inventory_excel() -> BytesIO:
    """Generate a two-sheet P&L Excel report using openpyxl for CFO insights."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    db = get_supabase()
    res = db.table("inventory").select("*").execute()
    items = res.data or []
    
    stock_value = sum((i.get("quantity", 0) or 0) * (i.get("price", 0.0) or 0.0) for i in items)
    total_revenue = sum((i.get("sold_quantity", 0) or 0) * (i.get("price", 0.0) or 0.0) for i in items)
    total_profit = sum(((i.get("price", 0.0) or 0.0) - (i.get("cost_price", 0.0) or 0.0)) * (i.get("sold_quantity", 0) or 0) for i in items)

    sorted_by_sales = sorted(items, key=lambda x: x.get("sold_quantity", 0) or 0, reverse=True)
    top_item = sorted_by_sales[0].get("name", "Нет") if sorted_by_sales else "Нет"

    ai_summary = _get_ai_summary(items, stock_value, total_revenue, total_profit, top_item)

    wb = Workbook()
    
    # ── Sheet 1: AI Аналитика ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "AI Аналитика"
    
    # Title
    ws1["A1"] = "Финансовый P&L Отчет и Метрики"
    ws1["A1"].font = Font(bold=True, size=16)
    
    # Metrics
    metrics = [
        ("Общая Выручка:", total_revenue),
        ("Общая Прибыль:", total_profit),
        ("Стоимость текущих остатков:", stock_value),
        ("Всего позиций на складе:", len(items))
    ]
    
    for idx, (label, val) in enumerate(metrics, start=2):
        ws1[f"A{idx}"] = label
        ws1[f"A{idx}"].font = Font(bold=True)
        ws1[f"B{idx}"] = val
        if idx <= 4:
            ws1[f"B{idx}"].number_format = '#,##0.00 "сом"'
    
    # AI Summary Block
    ws1.merge_cells('A7:F15')
    summary_cell = ws1["A7"]
    summary_cell.value = ai_summary
    summary_cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25

    # ── Sheet 2: Остатки Склада ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Остатки Склада")
    
    headers = [
        "Название товара", 
        "Текущий остаток на складе", 
        "Себестоимость (за 1 шт)", 
        "Цена продажи (за 1 шт)", 
        "Продано (шт)", 
        "Выручка", 
        "Прибыль"
    ]
    ws2.append(headers)
    
    # Style Header
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, column_title in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        
    # Append Data
    for i in items:
        name = i.get("name", "")
        qty = i.get("quantity", 0) or 0
        cost_price = i.get("cost_price", 0.0) or 0.0
        price = i.get("price", 0.0) or 0.0
        sold = i.get("sold_quantity", 0) or 0
        
        revenue = sold * price
        profit = (price - cost_price) * sold
        
        row = [
            name,
            qty,
            cost_price,
            price,
            sold,
            revenue,
            profit
        ]
        ws2.append(row)
        
    # Apply Filters
    max_row = len(items) + 1
    max_col = len(headers)
    ws2.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # Auto-adjust column widths & format numbers
    for col_idx, col in enumerate(ws2.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in col:
            try:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
            except:
                pass
                
            # Formatting financial columns
            if col_idx in [3, 4, 6, 7] and cell.row > 1:
                cell.number_format = '#,##0.00_-"сом"'
        
        # Add a bit more padding for header Russian words
        try:
            head_len = len(headers[col_idx-1])
        except IndexError:
            head_len = 0
        adjusted_width = float(max(int(max_length) + 2, int(head_len * 1.2)))
        ws2.column_dimensions[col_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
